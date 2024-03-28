
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Border, PatternFill, Alignment, Protection
import re

html_file_path = 'C:/Users/vladk/Desktop/lazy_parse/target.html'

with open(html_file_path, 'r', encoding='utf-8') as file:
    html_content = file.read()

soup = BeautifulSoup(html_content, 'html.parser')

table = soup.find('table')

excel_file_path = "готово.xlsx"

wb = openpyxl.load_workbook(excel_file_path)
source_sheet = wb['Микрометры']
category = soup.find('h1', class_='ty-mainbox-title')
name = category.text
base_name = "change name"
adder = 0
while True:
    try:
        ws = wb.create_sheet(name)
        break
    except ValueError:
        name = f"{base_name}_{adder}"
        adder += 1

item = 'Мерительный инструмент'
regex_pattern = re.compile(r'\b[A-ZА-Я][A-ZА-Я]+(?:[-][A-ZА-Я0-9]+)?\b')

category = category.find('span').text.strip()
description=''
th_tags = table.find_all('th', class_='grouped_title')
i = 0
number = 1
for th_tag in th_tags:
    p_tags = th_tag.find_all('p')
    description = ''
    for p_tag in p_tags:
        p_text = p_tag.get_text(strip=True) 
        description += p_text
    excluded_strings = set(text for p_tag in p_tags for text in p_tag.stripped_strings)
    th_text = ' '.join(text for text in th_tag.stripped_strings if text not in excluded_strings)
    subcategory = th_text
    subcategory = regex_pattern.sub('', th_text)
    subcategory = ' '.join(subcategory.split())
    subcategory = subcategory.strip()
    print(subcategory)   
    current_th_index = th_tags.index(th_tag)
    next_th_tag = th_tags[current_th_index + 1] if current_th_index + 1 < len(th_tags) else None
    rows_between_th = th_tag.find_all_next('tr')
    for row in rows_between_th:
        if 'header_group' in row.get('class', []):
            break
        if 'gr-txt' in row.get('class', []):
            name = row.find('a').text.strip()
            name = regex_pattern.sub('', name)
            name = re.sub(r'^[0-9A-ZА-Я]+[-][A-ZА-Я0-9.,]+','',name)
            name = re.sub(r'[,][ ]\d+(?:[\,\.\d]+)?[-]\d+(?:[\,\.\d]+)?[ʺ"]', '',name)
            name = re.sub(r'[,][ ]\d+(?:[\,\.\d]+)?[ʺ"][,]?', '',name)
            diapazon_izmereniy = row.find_all('td', class_='hidden_mobile')[2].text.strip()
            step_mm = row.find_all('td', class_='hidden_mobile')[3].text.strip()
            accuracy_mm = row.find_all('td', class_='hidden_mobile')[4].text.strip()
            have = 0
            if diapazon_izmereniy:
                diapazon_izmereniy = 'Диапозон измерений, мм: ' + str(diapazon_izmereniy)
                have = 1
            if step_mm:
                step_mm = '                                           Шаг, мм: ' + str(step_mm)
                have = 1
            if accuracy_mm:
                accuracy_mm = '                        Точность, мм: ' + str(accuracy_mm)
                have = 1
            specifications = diapazon_izmereniy + step_mm + accuracy_mm
            ws.append([number, name, item, category, subcategory, description, specifications])
            number +=1
            wb.save(excel_file_path)

            